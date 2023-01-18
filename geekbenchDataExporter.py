import ast
import json
from openpyxl import Workbook
#open gb5 file 
f = open("run4ARM.json")
data = json.load(f)
# print(data)
singleCoreTests = data["sections"][0]["workloads"]
multiCoreTests = data["sections"][1]["workloads"]
print(singleCoreTests)
#
wb = Workbook() 
ws = wb.active
ws.cell(row=1, column=1, value="Single Core Test")
ws.cell(row=1, column=2, value="Overall Score: " + str(data["sections"][0]["score"]))
for col in ws.iter_cols(min_row=2, max_col=1, max_row=21): 
    for cell in col: 
        temp = singleCoreTests.pop(0)
        cell.value =temp["name"]
        ws.cell(row=cell.row, column = cell.column+1, value = temp["score"])

ws.cell(row=1, column=3, value="Multi Core Test")
ws.cell(row=1, column=4, value="Overall Score: " + str(data["sections"][1]["score"]))
for col in ws.iter_cols(min_row=2, max_col=3, max_row=21, min_col=3): 
    for cell in col: 
        temp = multiCoreTests.pop(0)
        cell.value =temp["name"]
        ws.cell(row=cell.row, column = cell.column+1, value = temp["score"])



wb.save("run4ARM.xlsx")